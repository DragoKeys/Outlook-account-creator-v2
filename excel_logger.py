import os
from openpyxl import Workbook, load_workbook

HEADERS = [
    "Email",
    "Password",
    "First Name",
    "Last Name",
    "Birth Date",
    "Proxy Host",
    "Proxy Port",
    "Proxy Username",
    "Proxy Password",
    "Created At",
    "Col11",
    "Col12",
    "Col13",
    "Col14",
]

def append_account(row_data: list):
    file_path = "comptes.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        ws.append(HEADERS)
        ws.delete_rows(1)
    if len(row_data) < len(HEADERS):
        row_data = row_data + ["" for _ in range(len(HEADERS) - len(row_data))]
    else:
        row_data = row_data[:len(HEADERS)]
    ws.append(row_data)
    wb.save(file_path)
